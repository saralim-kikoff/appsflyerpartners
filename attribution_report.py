"""
AppsFlyer Monthly Attribution Report Generator

Pulls in-app events and Protect360 fraud data, applies flagging rules,
aggregates by agency, and outputs an Excel report + Slack notification.

Supports multiple apps:
- Kikoff: Credit line payments with Outside Attribution flagging
- Grant Cash Advance: First time offers with Add'l Fraud flagging
"""

import os
import requests
import pandas as pd
from datetime import datetime, timedelta
from io import StringIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import json


# =============================================================================
# CONFIGURATION
# =============================================================================

APPSFLYER_API_TOKEN = os.environ.get("APPSFLYER_API_TOKEN")
SLACK_WEBHOOK_URL = os.environ.get("SLACK_WEBHOOK_URL")

# -----------------------------------------------------------------------------
# KIKOFF APP CONFIGURATION
# -----------------------------------------------------------------------------
KIKOFF_APP_IDS = {
    "ios": "id1525159784",
    "android": "com.kikoff"
}
KIKOFF_EVENT_NAME = "CA Payment - Make Credit Line Success"

# Agencies authorized for view-through attribution (VTA)
VTA_AUTHORIZED_AGENCIES = ["adperiomedia", "globalwidemedia"]

# Attribution window limits
VTA_WINDOW_HOURS = 6
CTA_WINDOW_DAYS = 7

# -----------------------------------------------------------------------------
# GRANT CASH ADVANCE APP CONFIGURATION
# -----------------------------------------------------------------------------
GRANT_APP_IDS = {
    "ios": "id6472350114",
    "android": "com.kikoff.theseus"
}
GRANT_EVENT_NAME = "First Time Offer Accepted"

# Agencies to exclude from Grant data
GRANT_EXCLUDED_AGENCIES = ["mobiprobebd521", "unknown", ""]

BASE_URL = "https://hq1.appsflyer.com/api/raw-data/export/app"


# =============================================================================
# DATE HELPERS
# =============================================================================

def get_previous_month_range():
    """Get the first and last day of the previous month."""
    today = datetime.now()
    first_of_this_month = today.replace(day=1)
    last_of_prev_month = first_of_this_month - timedelta(days=1)
    first_of_prev_month = last_of_prev_month.replace(day=1)
    
    return (
        first_of_prev_month.strftime("%Y-%m-%d"),
        last_of_prev_month.strftime("%Y-%m-%d")
    )


def get_report_month_name():
    """Get the name of the previous month for report titles (e.g., 'November 2025')."""
    today = datetime.now()
    first_of_this_month = today.replace(day=1)
    last_of_prev_month = first_of_this_month - timedelta(days=1)
    return last_of_prev_month.strftime("%B %Y")


def get_report_month_yyyymm():
    """Get the previous month in YYYYMM format for filenames (e.g., '202511')."""
    today = datetime.now()
    first_of_this_month = today.replace(day=1)
    last_of_prev_month = first_of_this_month - timedelta(days=1)
    return last_of_prev_month.strftime("%Y%m")


# =============================================================================
# LOOKBACK PARSING
# =============================================================================

def parse_lookback_to_hours(value):
    """
    Convert AppsFlyer lookback format to hours.
    
    Examples:
        '7d' -> 168
        '6h' -> 6
        '24h' -> 24
        '30d' -> 720
    """
    if pd.isna(value) or value == '' or value is None:
        return 0
    
    value = str(value).strip().lower()
    
    try:
        if value.endswith('d'):
            return float(value[:-1]) * 24
        elif value.endswith('h'):
            return float(value[:-1])
        else:
            # Try to parse as numeric (assume hours)
            return float(value)
    except ValueError:
        return 0


# =============================================================================
# APPSFLYER API
# =============================================================================

def pull_appsflyer_report(app_id, report_type, from_date, to_date, event_name):
    """
    Pull raw data report from AppsFlyer Pull API.
    
    Args:
        app_id: iOS or Android app ID
        report_type: 'in_app_events' or 'protect360_in_app_events'
        from_date: Start date (YYYY-MM-DD)
        to_date: End date (YYYY-MM-DD)
        event_name: The event name to filter by
    
    Returns:
        pandas DataFrame with report data
    """
    # Map report type to endpoint
    endpoint_map = {
        "in_app_events": "in_app_events_report/v5",
        "protect360_in_app_events": "fraud-post-inapps/v5"
    }
    
    endpoint = endpoint_map.get(report_type)
    if not endpoint:
        raise ValueError(f"Unknown report type: {report_type}")
    
    url = f"{BASE_URL}/{app_id}/{endpoint}"
    
    headers = {
        "Authorization": f"Bearer {APPSFLYER_API_TOKEN}",
        "Accept": "text/csv"
    }
    
    params = {
        "from": from_date,
        "to": to_date,
        "event_name": event_name
    }
    
    print(f"Pulling {report_type} for {app_id}...")
    print(f"  URL: {url}")
    print(f"  Date range: {from_date} to {to_date}")
    
    response = requests.get(url, headers=headers, params=params)
    
    if response.status_code != 200:
        print(f"  Error: {response.status_code}")
        print(f"  Response: {response.text[:500] if response.text else 'Empty'}")
        return pd.DataFrame()
    
    if not response.text.strip():
        print("  No data returned")
        return pd.DataFrame()
    
    df = pd.read_csv(StringIO(response.text), low_memory=False)
    print(f"  Pulled {len(df)} rows")
    
    return df


def pull_all_reports(from_date, to_date, app_ids, event_name):
    """
    Pull all 4 reports (in-app events + P360 for iOS + Android).
    
    Args:
        from_date: Start date (YYYY-MM-DD)
        to_date: End date (YYYY-MM-DD)
        app_ids: Dict with 'ios' and 'android' app IDs
        event_name: The event name to filter by
    
    Returns:
        tuple: (delivered_df, fraud_df)
    """
    delivered_dfs = []
    fraud_dfs = []
    
    for platform, app_id in app_ids.items():
        # In-app events (delivered)
        df = pull_appsflyer_report(app_id, "in_app_events", from_date, to_date, event_name)
        if not df.empty:
            df["platform"] = platform
            delivered_dfs.append(df)
        
        # Protect360 fraud events
        df = pull_appsflyer_report(app_id, "protect360_in_app_events", from_date, to_date, event_name)
        if not df.empty:
            df["platform"] = platform
            fraud_dfs.append(df)
    
    delivered_df = pd.concat(delivered_dfs, ignore_index=True) if delivered_dfs else pd.DataFrame()
    fraud_df = pd.concat(fraud_dfs, ignore_index=True) if fraud_dfs else pd.DataFrame()
    
    # Filter out organic events (non-organic only)
    if not delivered_df.empty:
        delivered_df.columns = delivered_df.columns.str.strip().str.lower().str.replace(' ', '_')
        if 'media_source' in delivered_df.columns:
            delivered_df = delivered_df[delivered_df['media_source'].str.lower() != 'organic']
            print(f"After filtering organic: {len(delivered_df)} delivered events")
    
    if not fraud_df.empty:
        fraud_df.columns = fraud_df.columns.str.strip().str.lower().str.replace(' ', '_')
        if 'media_source' in fraud_df.columns:
            fraud_df = fraud_df[fraud_df['media_source'].str.lower() != 'organic']
            print(f"After filtering organic: {len(fraud_df)} fraud events")
    
    return delivered_df, fraud_df


# =============================================================================
# FLAGGING LOGIC
# =============================================================================

def apply_kikoff_flagging_rules(df):
    """
    Apply Kikoff-specific flagging rules to identify suspicious events (Outside Attribution).
    
    Rules:
    1. Unauthorized VTA: impression attribution from non-authorized agencies
    2. VTA Window Exceeded: impression from authorized agencies > 6h lookback
    3. CTA Window Exceeded: click attribution > 7 days lookback
    
    Returns:
        DataFrame with 'is_flagged' and 'flag_reason' columns added
    """
    if df.empty:
        df["is_flagged"] = False
        df["flag_reason"] = ""
        return df
    
    # Normalize column names (AppsFlyer uses various formats)
    df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
    
    # Identify the agency column (could be 'agency', 'partner', 'media_source', etc.)
    agency_col = None
    for col in ['agency', 'partner', 'af_prt', 'media_source']:
        if col in df.columns:
            agency_col = col
            break
    
    if agency_col is None:
        print("Warning: Could not find agency column. Using 'unknown'.")
        df["agency_normalized"] = "unknown"
    else:
        df["agency_normalized"] = df[agency_col].fillna("unknown").str.strip().str.lower()
    
    # Identify touch type column
    touch_type_col = None
    for col in ['attributed_touch_type', 'touch_type']:
        if col in df.columns:
            touch_type_col = col
            break
    
    if touch_type_col is None:
        print("Warning: Could not find touch type column.")
        df["touch_type_normalized"] = "unknown"
    else:
        df["touch_type_normalized"] = df[touch_type_col].fillna("").str.strip().str.lower()
    
    # Identify lookback column
    lookback_col = None
    for col in ['attribution_lookback', 'lookback', 'time_to_install']:
        if col in df.columns:
            lookback_col = col
            break
    
    if lookback_col:
        df["lookback_hours"] = df[lookback_col].apply(parse_lookback_to_hours)
    else:
        print("Warning: Could not find lookback column. Setting to 0.")
        df["lookback_hours"] = 0
    
    # Apply flagging rules
    df["is_flagged"] = False
    df["flag_reason"] = ""
    
    # Rule 1: Unauthorized VTA
    mask_unauthorized_vta = (
        (df["touch_type_normalized"] == "impression") &
        (~df["agency_normalized"].isin(VTA_AUTHORIZED_AGENCIES))
    )
    df.loc[mask_unauthorized_vta, "is_flagged"] = True
    df.loc[mask_unauthorized_vta, "flag_reason"] = "Unauthorized VTA"
    
    # Rule 2: VTA Window Exceeded (for authorized agencies)
    mask_vta_window = (
        (df["touch_type_normalized"] == "impression") &
        (df["agency_normalized"].isin(VTA_AUTHORIZED_AGENCIES)) &
        (df["lookback_hours"] > VTA_WINDOW_HOURS)
    )
    df.loc[mask_vta_window, "is_flagged"] = True
    df.loc[mask_vta_window, "flag_reason"] = "VTA Window Exceeded (>6h)"
    
    # Rule 3: CTA Window Exceeded
    mask_cta_window = (
        (df["touch_type_normalized"] == "click") &
        (df["lookback_hours"] > CTA_WINDOW_DAYS * 24)
    )
    df.loc[mask_cta_window, "is_flagged"] = True
    df.loc[mask_cta_window, "flag_reason"] = "CTA Window Exceeded (>7d)"
    
    flagged_count = df["is_flagged"].sum()
    print(f"Flagged {flagged_count} events out of {len(df)}")
    
    return df


def apply_grant_addl_fraud_rules(df, fraud_df):
    """
    Apply Grant-specific flagging rules to identify Add'l Fraud events.
    
    Rules:
    1. Event Value does NOT contain "00}"
    2. Not already in P360 fraud data (matched by AppsFlyer ID and Customer ID)
    
    Returns:
        DataFrame of flagged events (Add'l Fraud)
    """
    if df.empty:
        return pd.DataFrame()
    
    # Normalize column names
    df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
    
    # Find event_value column
    event_value_col = None
    for col in ['event_value', 'event_revenue', 'revenue']:
        if col in df.columns:
            event_value_col = col
            break
    
    if event_value_col is None:
        print("Warning: Could not find event_value column for Grant Add'l Fraud check.")
        return pd.DataFrame()
    
    # Rule 1: Event Value does NOT contain "00}"
    # Convert to string and check
    df["_event_value_str"] = df[event_value_col].fillna("").astype(str)
    addl_fraud_mask = ~df["_event_value_str"].str.contains("00}", regex=False)
    
    addl_fraud_df = df[addl_fraud_mask].copy()
    print(f"Events with event_value not containing '00}}': {len(addl_fraud_df)}")
    
    if addl_fraud_df.empty:
        return pd.DataFrame()
    
    # Rule 2: Not already in P360 fraud data
    if not fraud_df.empty:
        fraud_df.columns = fraud_df.columns.str.strip().str.lower().str.replace(' ', '_')
        
        # Find customer_id and appsflyer_id columns
        customer_id_col = None
        appsflyer_id_col = None
        
        for col in addl_fraud_df.columns:
            if 'customer' in col and 'id' in col:
                customer_id_col = col
            if 'appsflyer' in col and 'id' in col:
                appsflyer_id_col = col
        
        if customer_id_col and appsflyer_id_col:
            # Create composite key for matching
            addl_fraud_df["_match_key"] = (
                addl_fraud_df[customer_id_col].astype(str) + "_" + 
                addl_fraud_df[appsflyer_id_col].astype(str)
            )
            fraud_df["_match_key"] = (
                fraud_df[customer_id_col].astype(str) + "_" + 
                fraud_df[appsflyer_id_col].astype(str)
            )
            
            # Remove events that exist in fraud data
            fraud_keys = set(fraud_df["_match_key"].unique())
            addl_fraud_df = addl_fraud_df[~addl_fraud_df["_match_key"].isin(fraud_keys)]
            
            # Clean up temp columns
            addl_fraud_df = addl_fraud_df.drop(columns=["_match_key", "_event_value_str"], errors='ignore')
            
            print(f"Add'l Fraud events (after removing P360 matches): {len(addl_fraud_df)}")
        else:
            print("Warning: Could not find Customer ID or AppsFlyer ID columns for deduplication")
            addl_fraud_df = addl_fraud_df.drop(columns=["_event_value_str"], errors='ignore')
    else:
        addl_fraud_df = addl_fraud_df.drop(columns=["_event_value_str"], errors='ignore')
    
    # Normalize agency column for aggregation
    agency_col = None
    for col in ['agency', 'partner', 'af_prt', 'media_source']:
        if col in addl_fraud_df.columns:
            agency_col = col
            break
    
    if agency_col:
        addl_fraud_df["agency_normalized"] = addl_fraud_df[agency_col].fillna("unknown").str.strip().str.lower()
    else:
        addl_fraud_df["agency_normalized"] = "unknown"
    
    return addl_fraud_df


# =============================================================================
# AGGREGATION
# =============================================================================

def aggregate_by_agency(df, value_col_name="event_count"):
    """
    Aggregate event counts by agency.
    Excludes 'unknown' agency from aggregations.
    
    Returns:
        DataFrame with agency and event count
    """
    if df.empty:
        return pd.DataFrame(columns=["agency", value_col_name])
    
    # Ensure agency column exists
    if "agency_normalized" not in df.columns:
        # Try to find and normalize agency column
        agency_col = None
        for col in ['agency', 'partner', 'af_prt', 'media_source']:
            if col in df.columns:
                agency_col = col
                break
        
        if agency_col:
            df["agency_normalized"] = df[agency_col].fillna("unknown").str.strip().str.lower()
        else:
            df["agency_normalized"] = "unknown"
    
    # Filter out unknown agencies
    df_filtered = df[df["agency_normalized"] != "unknown"]
    
    if df_filtered.empty:
        return pd.DataFrame(columns=["agency", value_col_name])
    
    aggregated = df_filtered.groupby("agency_normalized").size().reset_index(name=value_col_name)
    aggregated = aggregated.rename(columns={"agency_normalized": "agency"})
    aggregated = aggregated.sort_values(value_col_name, ascending=False)
    
    return aggregated


# =============================================================================
# EXCEL GENERATION
# =============================================================================

def style_header(ws, row_num, num_cols):
    """Apply header styling to a row."""
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")


def add_dataframe_to_sheet(ws, df, start_row=1):
    """Add a DataFrame to a worksheet with formatting."""
    if df.empty:
        ws.cell(row=start_row, column=1, value="No data")
        return
    
    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        header_text = str(col_name).replace("_", " ").title()
        ws.cell(row=start_row, column=col_idx, value=header_text)
    
    style_header(ws, start_row, len(df.columns))
    
    # Write data
    for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center")
    
    # Auto-adjust column widths
    from openpyxl.utils import get_column_letter
    for col_idx, col_name in enumerate(df.columns, 1):
        try:
            # Get max length of header
            header_len = len(str(col_name))
            # Get max length of data in column (using iloc for safer access)
            if len(df) > 0:
                data_len = df.iloc[:, col_idx - 1].astype(str).str.len().max()
            else:
                data_len = 0
            max_length = max(header_len, data_len)
            # Cap width at 50 to avoid extremely wide columns
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
        except Exception:
            # Default width if calculation fails
            ws.column_dimensions[get_column_letter(col_idx)].width = 15


def generate_excel_report(summary_df, delivered_df, fraud_df, outside_attr_df, report_month):
    """
    Generate Excel workbook with multiple tabs.
    
    Tabs:
    1. Summary - Agency rollup with Net Valid calculation
    2. Delivered Events - Raw event data
    3. Fraud Events - Raw P360 event data
    4. Outside Attribution Events - Raw flagged event data
    """
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Reorder summary columns to put net_valid last
    if not summary_df.empty:
        # Define desired column order
        desired_order = ["agency", "delivered", "fraud", "fraud_rate_%", "outside_attribution", "outside_attr_rate_%", "net_valid"]
        # Only include columns that exist
        final_order = [col for col in desired_order if col in summary_df.columns]
        # Add any remaining columns not in the desired order
        remaining = [col for col in summary_df.columns if col not in final_order]
        final_order.extend(remaining)
        summary_df = summary_df[final_order]
    
    # Tab 1: Summary
    ws_summary = wb.create_sheet("Summary")
    ws_summary.cell(row=1, column=1, value=f"Attribution Report - {report_month}")
    ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_summary.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    add_dataframe_to_sheet(ws_summary, summary_df, start_row=3)
    
    # Tab 2: Delivered Events (raw data)
    ws_delivered = wb.create_sheet("Delivered Events")
    add_dataframe_to_sheet(ws_delivered, delivered_df)
    
    # Tab 3: Fraud Events (raw data)
    ws_fraud = wb.create_sheet("Fraud Events")
    add_dataframe_to_sheet(ws_fraud, fraud_df)
    
    # Tab 4: Outside Attribution Events (raw data)
    ws_outside_attr = wb.create_sheet("Outside Attribution Events")
    add_dataframe_to_sheet(ws_outside_attr, outside_attr_df)
    
    # Save workbook
    filename = f"attribution_report_{report_month.replace(' ', '_').lower()}.xlsx"
    filepath = f"/tmp/{filename}"
    wb.save(filepath)
    
    print(f"Excel report saved: {filepath}")
    return filepath


# =============================================================================
# SLACK NOTIFICATION
# =============================================================================

SLACK_BOT_TOKEN = os.environ.get("SLACK_BOT_TOKEN")
SLACK_CHANNEL_ID = os.environ.get("SLACK_CHANNEL_ID")


def upload_file_to_slack(filepath, channel_id, initial_comment=""):
    """
    Upload a file to Slack using the Bot Token.
    Returns the file permalink if successful, None otherwise.
    """
    if not SLACK_BOT_TOKEN or not channel_id:
        return None
    
    url = "https://slack.com/api/files.upload"
    
    headers = {
        "Authorization": f"Bearer {SLACK_BOT_TOKEN}"
    }
    
    with open(filepath, "rb") as f:
        response = requests.post(
            url,
            headers=headers,
            data={
                "channels": channel_id,
                "initial_comment": initial_comment,
                "filename": os.path.basename(filepath),
                "title": os.path.basename(filepath)
            },
            files={"file": f}
        )
    
    if response.status_code == 200:
        result = response.json()
        if result.get("ok"):
            permalink = result.get("file", {}).get("permalink", "")
            print(f"File uploaded to Slack successfully: {permalink}")
            return permalink
        else:
            print(f"Slack file upload failed: {result.get('error')}")
    else:
        print(f"Slack file upload failed: {response.status_code}")
    
    return None


def send_slack_notification(summary_df, report_month, excel_filepath):
    """
    Send Slack message with report summary.
    
    If SLACK_BOT_TOKEN and SLACK_CHANNEL_ID are set, uploads the Excel file directly.
    Otherwise, falls back to webhook with GitHub Actions artifact link.
    """
    # Try file upload first if bot token is configured
    file_permalink = None
    if SLACK_BOT_TOKEN and SLACK_CHANNEL_ID and excel_filepath:
        file_permalink = upload_file_to_slack(
            excel_filepath, 
            SLACK_CHANNEL_ID,
            initial_comment=""  # We'll send the summary separately
        )
    
    # Build the summary message
    if not SLACK_WEBHOOK_URL and not SLACK_BOT_TOKEN:
        print("No Slack credentials configured. Skipping notification.")
        return
    
    # Handle empty data case
    if summary_df.empty:
        blocks = [
            {
                "type": "header",
                "text": {
                    "type": "plain_text",
                    "text": f"📊 Monthly AppsFlyer Partner Report — {report_month}",
                    "emoji": True
                }
            },
            {
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": "⚠️ *No data found for this period.*\n\nThis could mean:\n• No events matched the criteria\n• API permissions issue\n• Event name mismatch"
                }
            }
        ]
        
        message = {"blocks": blocks}
    else:
        # Build summary table
        total_delivered = int(summary_df["delivered"].sum()) if "delivered" in summary_df.columns else 0
        total_fraud = int(summary_df["fraud"].sum()) if "fraud" in summary_df.columns else 0
        total_outside_attr = int(summary_df["outside_attribution"].sum()) if "outside_attribution" in summary_df.columns else 0
        total_net_valid = int(summary_df["net_valid"].sum()) if "net_valid" in summary_df.columns else 0
        
        overall_fraud_rate = (total_fraud / total_delivered * 100) if total_delivered > 0 else 0
        overall_outside_attr_rate = (total_outside_attr / total_delivered * 100) if total_delivered > 0 else 0
        
        # Top agencies by net valid
        top_agencies = summary_df.head(5).to_dict('records') if not summary_df.empty else []
        
        agency_lines = []
        for agency in top_agencies:
            name = agency.get("agency", "Unknown")
            delivered = int(agency.get("delivered", 0))
            net_valid = int(agency.get("net_valid", 0))
            fraud_rate = float(agency.get("fraud_rate_%", 0))
            agency_lines.append(f"• *{name}*: {net_valid:,} net valid / {delivered:,} delivered ({fraud_rate:.1f}% fraud)")
        
        agency_summary = "\n".join(agency_lines) if agency_lines else "No data"
        
        blocks = [
            {
                "type": "header",
                "text": {
                    "type": "plain_text",
                    "text": f"📊 Monthly AppsFlyer Partner Report — {report_month}",
                    "emoji": True
                }
            },
            {
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": f"*Overall Totals*\n"
                            f"• Delivered Events: *{total_delivered:,}*\n"
                            f"• Fraud Events (P360): *{total_fraud:,}* ({overall_fraud_rate:.1f}%)\n"
                            f"• Outside Attribution Events: *{total_outside_attr:,}* ({overall_outside_attr_rate:.1f}%)\n"
                            f"• Net Valid Events: *{total_net_valid:,}*"
                }
            },
            {
                "type": "divider"
            },
            {
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": f"*Top Agencies by Net Valid*\n{agency_summary}"
                }
            },
            {
                "type": "context",
                "elements": [
                    {
                        "type": "mrkdwn",
                        "text": f"Event: `{EVENT_NAME}`"
                    }
                ]
            }
        ]
        
        # Add download link
        if file_permalink:
            # File was uploaded to Slack - link to it
            blocks.append({
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": f"📎 <{file_permalink}|Download Full Report>"
                }
            })
        else:
            # Fall back to GitHub Actions artifact link
            github_repo = os.environ.get("GITHUB_REPOSITORY", "")
            github_run_id = os.environ.get("GITHUB_RUN_ID", "")
            
            if github_repo and github_run_id:
                artifact_url = f"https://github.com/{github_repo}/actions/runs/{github_run_id}"
                blocks.append({
                    "type": "section",
                    "text": {
                        "type": "mrkdwn",
                        "text": f"📎 <{artifact_url}|Download Full Report> (requires GitHub access)"
                    }
                })
        
        message = {"blocks": blocks}
        
        message = {"blocks": blocks}
    
    response = requests.post(
        SLACK_WEBHOOK_URL,
        json=message,
        headers={"Content-Type": "application/json"}
    )
    
    if response.status_code == 200:
        print("Slack notification sent successfully")
    else:
        print(f"Slack notification failed: {response.status_code} - {response.text}")


# =============================================================================
# MAIN EXECUTION
# =============================================================================

def process_kikoff_app(from_date, to_date):
    """Process Kikoff app data and return results."""
    print("\n" + "=" * 60)
    print("KIKOFF APP")
    print("=" * 60)
    
    # Pull all reports
    print("\n" + "-" * 40)
    print("Pulling AppsFlyer Data")
    print("-" * 40)
    
    delivered_df, fraud_df = pull_all_reports(from_date, to_date, KIKOFF_APP_IDS, KIKOFF_EVENT_NAME)
    
    print(f"\nTotal delivered events: {len(delivered_df)}")
    print(f"Total fraud events: {len(fraud_df)}")
    
    # Apply flagging rules to delivered events
    print("\n" + "-" * 40)
    print("Applying Outside Attribution Rules")
    print("-" * 40)
    
    delivered_df = apply_kikoff_flagging_rules(delivered_df)
    
    # Separate flagged events
    if not delivered_df.empty and "is_flagged" in delivered_df.columns:
        flagged_events_df = delivered_df[delivered_df["is_flagged"] == True].copy()
    else:
        flagged_events_df = pd.DataFrame()
    
    print(f"Outside attribution events (before dedup): {len(flagged_events_df)}")
    
    # Remove outside attribution events that are already in P360 fraud data
    if not flagged_events_df.empty and not fraud_df.empty:
        flagged_events_df.columns = flagged_events_df.columns.str.strip().str.lower().str.replace(' ', '_')
        fraud_df.columns = fraud_df.columns.str.strip().str.lower().str.replace(' ', '_')
        
        customer_id_col = None
        appsflyer_id_col = None
        
        for col in flagged_events_df.columns:
            if 'customer' in col and 'id' in col:
                customer_id_col = col
            if 'appsflyer' in col and 'id' in col:
                appsflyer_id_col = col
        
        if customer_id_col and appsflyer_id_col:
            flagged_events_df["_match_key"] = (
                flagged_events_df[customer_id_col].astype(str) + "_" + 
                flagged_events_df[appsflyer_id_col].astype(str)
            )
            fraud_df["_match_key"] = (
                fraud_df[customer_id_col].astype(str) + "_" + 
                fraud_df[appsflyer_id_col].astype(str)
            )
            
            fraud_keys = set(fraud_df["_match_key"].unique())
            flagged_events_df = flagged_events_df[~flagged_events_df["_match_key"].isin(fraud_keys)]
            
            flagged_events_df = flagged_events_df.drop(columns=["_match_key"])
            fraud_df = fraud_df.drop(columns=["_match_key"])
            
            print(f"Outside attribution events (after dedup): {len(flagged_events_df)}")
    
    # Aggregate by agency
    print("\n" + "-" * 40)
    print("Aggregating by Agency")
    print("-" * 40)
    
    delivered_agg = aggregate_by_agency(delivered_df, "delivered")
    fraud_agg = aggregate_by_agency(fraud_df, "fraud")
    flagged_agg = aggregate_by_agency(flagged_events_df, "outside_attribution")
    
    # Build summary
    if delivered_agg.empty:
        summary_df = pd.DataFrame(columns=["agency", "delivered", "fraud", "outside_attribution", "fraud_rate_%", "outside_attr_rate_%", "net_valid"])
    else:
        summary_df = delivered_agg.copy()
        
        if not fraud_agg.empty:
            summary_df = summary_df.merge(fraud_agg, on="agency", how="left")
        else:
            summary_df["fraud"] = 0
            
        if not flagged_agg.empty:
            summary_df = summary_df.merge(flagged_agg[["agency", "outside_attribution"]], on="agency", how="left")
        else:
            summary_df["outside_attribution"] = 0
        
        summary_df = summary_df.fillna(0)
        
        for col in ["delivered", "fraud", "outside_attribution"]:
            if col in summary_df.columns:
                summary_df[col] = pd.to_numeric(summary_df[col], errors='coerce').fillna(0)
        
        summary_df["net_valid"] = (
            summary_df["delivered"] - summary_df["fraud"] - summary_df["outside_attribution"]
        ).astype(int)
        
        summary_df["fraud_rate_%"] = summary_df.apply(
            lambda row: round(row["fraud"] / row["delivered"] * 100, 1) if row["delivered"] > 0 else 0,
            axis=1
        )
        summary_df["outside_attr_rate_%"] = summary_df.apply(
            lambda row: round(row["outside_attribution"] / row["delivered"] * 100, 1) if row["delivered"] > 0 else 0,
            axis=1
        )
        
        for col in ["delivered", "fraud", "outside_attribution", "net_valid"]:
            if col in summary_df.columns:
                summary_df[col] = summary_df[col].astype(int)
        
        summary_df = summary_df.sort_values("delivered", ascending=False)
    
    print("\nKikoff Summary:")
    if not summary_df.empty:
        print(summary_df.to_string(index=False))
    
    return {
        "summary": summary_df,
        "delivered": delivered_df,
        "fraud": fraud_df,
        "flagged": flagged_events_df,
        "flagged_label": "Outside Attribution Events"
    }


def process_grant_app(from_date, to_date):
    """Process Grant Cash Advance app data and return results."""
    print("\n" + "=" * 60)
    print("GRANT CASH ADVANCE APP")
    print("=" * 60)
    
    # Pull all reports
    print("\n" + "-" * 40)
    print("Pulling AppsFlyer Data")
    print("-" * 40)
    
    delivered_df, fraud_df = pull_all_reports(from_date, to_date, GRANT_APP_IDS, GRANT_EVENT_NAME)
    
    print(f"\nTotal delivered events: {len(delivered_df)}")
    print(f"Total fraud events: {len(fraud_df)}")
    
    # Apply Add'l Fraud rules
    print("\n" + "-" * 40)
    print("Applying Add'l Fraud Rules")
    print("-" * 40)
    
    addl_fraud_df = apply_grant_addl_fraud_rules(delivered_df, fraud_df)
    
    print(f"Add'l Fraud events: {len(addl_fraud_df)}")
    
    # Normalize delivered_df columns for aggregation
    if not delivered_df.empty:
        delivered_df.columns = delivered_df.columns.str.strip().str.lower().str.replace(' ', '_')
        agency_col = None
        for col in ['agency', 'partner', 'af_prt', 'media_source']:
            if col in delivered_df.columns:
                agency_col = col
                break
        if agency_col:
            delivered_df["agency_normalized"] = delivered_df[agency_col].fillna("unknown").str.strip().str.lower()
        else:
            delivered_df["agency_normalized"] = "unknown"
        
        # Filter out excluded agencies
        before_count = len(delivered_df)
        delivered_df = delivered_df[~delivered_df["agency_normalized"].isin(GRANT_EXCLUDED_AGENCIES)]
        delivered_df = delivered_df[delivered_df["agency_normalized"].notna()]
        print(f"Delivered events after filtering excluded agencies: {len(delivered_df)} (removed {before_count - len(delivered_df)})")
    
    # Normalize fraud_df columns for aggregation
    if not fraud_df.empty:
        fraud_df.columns = fraud_df.columns.str.strip().str.lower().str.replace(' ', '_')
        agency_col = None
        for col in ['agency', 'partner', 'af_prt', 'media_source']:
            if col in fraud_df.columns:
                agency_col = col
                break
        if agency_col:
            fraud_df["agency_normalized"] = fraud_df[agency_col].fillna("unknown").str.strip().str.lower()
        else:
            fraud_df["agency_normalized"] = "unknown"
        
        # Filter out excluded agencies
        before_count = len(fraud_df)
        fraud_df = fraud_df[~fraud_df["agency_normalized"].isin(GRANT_EXCLUDED_AGENCIES)]
        fraud_df = fraud_df[fraud_df["agency_normalized"].notna()]
        print(f"Fraud events after filtering excluded agencies: {len(fraud_df)} (removed {before_count - len(fraud_df)})")
    
    # Filter excluded agencies from addl_fraud_df as well
    if not addl_fraud_df.empty:
        before_count = len(addl_fraud_df)
        addl_fraud_df = addl_fraud_df[~addl_fraud_df["agency_normalized"].isin(GRANT_EXCLUDED_AGENCIES)]
        addl_fraud_df = addl_fraud_df[addl_fraud_df["agency_normalized"].notna()]
        print(f"Add'l Fraud events after filtering excluded agencies: {len(addl_fraud_df)} (removed {before_count - len(addl_fraud_df)})")
    
    # Aggregate by agency
    print("\n" + "-" * 40)
    print("Aggregating by Agency")
    print("-" * 40)
    
    delivered_agg = aggregate_by_agency(delivered_df, "delivered")
    fraud_agg = aggregate_by_agency(fraud_df, "fraud")
    addl_fraud_agg = aggregate_by_agency(addl_fraud_df, "addl_fraud")
    
    # Build summary
    if delivered_agg.empty:
        summary_df = pd.DataFrame(columns=["agency", "delivered", "fraud", "addl_fraud", "fraud_rate_%", "addl_fraud_rate_%", "net_valid"])
    else:
        summary_df = delivered_agg.copy()
        
        if not fraud_agg.empty:
            summary_df = summary_df.merge(fraud_agg, on="agency", how="left")
        else:
            summary_df["fraud"] = 0
            
        if not addl_fraud_agg.empty:
            summary_df = summary_df.merge(addl_fraud_agg[["agency", "addl_fraud"]], on="agency", how="left")
        else:
            summary_df["addl_fraud"] = 0
        
        summary_df = summary_df.fillna(0)
        
        for col in ["delivered", "fraud", "addl_fraud"]:
            if col in summary_df.columns:
                summary_df[col] = pd.to_numeric(summary_df[col], errors='coerce').fillna(0)
        
        summary_df["net_valid"] = (
            summary_df["delivered"] - summary_df["fraud"] - summary_df["addl_fraud"]
        ).astype(int)
        
        summary_df["fraud_rate_%"] = summary_df.apply(
            lambda row: round(row["fraud"] / row["delivered"] * 100, 1) if row["delivered"] > 0 else 0,
            axis=1
        )
        summary_df["addl_fraud_rate_%"] = summary_df.apply(
            lambda row: round(row["addl_fraud"] / row["delivered"] * 100, 1) if row["delivered"] > 0 else 0,
            axis=1
        )
        
        for col in ["delivered", "fraud", "addl_fraud", "net_valid"]:
            if col in summary_df.columns:
                summary_df[col] = summary_df[col].astype(int)
        
        summary_df = summary_df.sort_values("delivered", ascending=False)
    
    print("\nGrant Summary:")
    if not summary_df.empty:
        print(summary_df.to_string(index=False))
    
    return {
        "summary": summary_df,
        "delivered": delivered_df,
        "fraud": fraud_df,
        "flagged": addl_fraud_df,
        "flagged_label": "Add'l Fraud Events"
    }


def generate_kikoff_excel_report(kikoff_data, report_month):
    """
    Generate Excel workbook for Kikoff app.
    """
    wb = Workbook()
    wb.remove(wb.active)
    
    # Reorder Kikoff summary columns
    kikoff_summary = kikoff_data["summary"].copy() if not kikoff_data["summary"].empty else pd.DataFrame()
    if not kikoff_summary.empty:
        desired_order = ["agency", "delivered", "fraud", "fraud_rate_%", "outside_attribution", "outside_attr_rate_%", "net_valid"]
        final_order = [col for col in desired_order if col in kikoff_summary.columns]
        remaining = [col for col in kikoff_summary.columns if col not in final_order]
        final_order.extend(remaining)
        kikoff_summary = kikoff_summary[final_order]
    
    ws = wb.create_sheet("Summary")
    ws.cell(row=1, column=1, value=f"Kikoff Attribution Report - {report_month}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    add_dataframe_to_sheet(ws, kikoff_summary, start_row=3)
    
    ws = wb.create_sheet("Delivered Events")
    add_dataframe_to_sheet(ws, kikoff_data["delivered"])
    
    ws = wb.create_sheet("Fraud Events")
    add_dataframe_to_sheet(ws, kikoff_data["fraud"])
    
    ws = wb.create_sheet("Outside Attribution Events")
    add_dataframe_to_sheet(ws, kikoff_data["flagged"])
    
    # Save workbook
    month_yyyymm = get_report_month_yyyymm()
    filename = f"Appsflyer_Kikoff_{month_yyyymm}.xlsx"
    filepath = f"/tmp/{filename}"
    wb.save(filepath)
    
    print(f"Kikoff Excel report saved: {filepath}")
    return filepath


def generate_grant_excel_report(grant_data, report_month):
    """
    Generate Excel workbook for Grant Cash Advance app.
    """
    wb = Workbook()
    wb.remove(wb.active)
    
    # Reorder Grant summary columns
    grant_summary = grant_data["summary"].copy() if not grant_data["summary"].empty else pd.DataFrame()
    if not grant_summary.empty:
        desired_order = ["agency", "delivered", "fraud", "fraud_rate_%", "addl_fraud", "addl_fraud_rate_%", "net_valid"]
        final_order = [col for col in desired_order if col in grant_summary.columns]
        remaining = [col for col in grant_summary.columns if col not in final_order]
        final_order.extend(remaining)
        grant_summary = grant_summary[final_order]
    
    ws = wb.create_sheet("Summary")
    ws.cell(row=1, column=1, value=f"Grant Cash Advance Attribution Report - {report_month}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    add_dataframe_to_sheet(ws, grant_summary, start_row=3)
    
    ws = wb.create_sheet("Delivered Events")
    add_dataframe_to_sheet(ws, grant_data["delivered"])
    
    ws = wb.create_sheet("Fraud Events")
    add_dataframe_to_sheet(ws, grant_data["fraud"])
    
    ws = wb.create_sheet("Add'l Fraud Events")
    add_dataframe_to_sheet(ws, grant_data["flagged"])
    
    # Save workbook
    month_yyyymm = get_report_month_yyyymm()
    filename = f"Appsflyer_Grant_{month_yyyymm}.xlsx"
    filepath = f"/tmp/{filename}"
    wb.save(filepath)
    
    print(f"Grant Excel report saved: {filepath}")
    return filepath


def send_combined_slack_notification(kikoff_data, grant_data, report_month, kikoff_filepath, grant_filepath):
    """
    Send Slack message with combined report summary for both apps.
    """
    # Try file upload first if bot token is configured
    kikoff_permalink = None
    grant_permalink = None
    
    if SLACK_BOT_TOKEN and SLACK_CHANNEL_ID:
        if kikoff_filepath:
            kikoff_permalink = upload_file_to_slack(
                kikoff_filepath, 
                SLACK_CHANNEL_ID,
                initial_comment=""
            )
        if grant_filepath:
            grant_permalink = upload_file_to_slack(
                grant_filepath, 
                SLACK_CHANNEL_ID,
                initial_comment=""
            )
    
    if not SLACK_WEBHOOK_URL and not SLACK_BOT_TOKEN:
        print("No Slack credentials configured. Skipping notification.")
        return
    
    # Build Kikoff summary
    kikoff_summary = kikoff_data["summary"]
    if not kikoff_summary.empty:
        kikoff_delivered = int(kikoff_summary["delivered"].sum())
        kikoff_fraud = int(kikoff_summary["fraud"].sum())
        kikoff_outside_attr = int(kikoff_summary["outside_attribution"].sum()) if "outside_attribution" in kikoff_summary.columns else 0
        kikoff_net_valid = int(kikoff_summary["net_valid"].sum())
        kikoff_fraud_rate = (kikoff_fraud / kikoff_delivered * 100) if kikoff_delivered > 0 else 0
        kikoff_outside_attr_rate = (kikoff_outside_attr / kikoff_delivered * 100) if kikoff_delivered > 0 else 0
    else:
        kikoff_delivered = kikoff_fraud = kikoff_outside_attr = kikoff_net_valid = 0
        kikoff_fraud_rate = kikoff_outside_attr_rate = 0
    
    # Build Grant summary
    grant_summary = grant_data["summary"]
    if not grant_summary.empty:
        grant_delivered = int(grant_summary["delivered"].sum())
        grant_fraud = int(grant_summary["fraud"].sum())
        grant_addl_fraud = int(grant_summary["addl_fraud"].sum()) if "addl_fraud" in grant_summary.columns else 0
        grant_net_valid = int(grant_summary["net_valid"].sum())
        grant_fraud_rate = (grant_fraud / grant_delivered * 100) if grant_delivered > 0 else 0
        grant_addl_fraud_rate = (grant_addl_fraud / grant_delivered * 100) if grant_delivered > 0 else 0
    else:
        grant_delivered = grant_fraud = grant_addl_fraud = grant_net_valid = 0
        grant_fraud_rate = grant_addl_fraud_rate = 0
    
    blocks = [
        {
            "type": "header",
            "text": {
                "type": "plain_text",
                "text": f"📊 Monthly AppsFlyer Partner Report — {report_month}",
                "emoji": True
            }
        },
        {
            "type": "section",
            "text": {
                "type": "mrkdwn",
                "text": f"*KIKOFF* (`{KIKOFF_EVENT_NAME}`)\n"
                        f"• Delivered: *{kikoff_delivered:,}*\n"
                        f"• Fraud (P360): *{kikoff_fraud:,}* ({kikoff_fraud_rate:.1f}%)\n"
                        f"• Outside Attribution: *{kikoff_outside_attr:,}* ({kikoff_outside_attr_rate:.1f}%)\n"
                        f"• Net Valid: *{kikoff_net_valid:,}*"
            }
        },
        {
            "type": "divider"
        },
        {
            "type": "section",
            "text": {
                "type": "mrkdwn",
                "text": f"*GRANT CASH ADVANCE* (`{GRANT_EVENT_NAME}`)\n"
                        f"• Delivered: *{grant_delivered:,}*\n"
                        f"• Fraud (P360): *{grant_fraud:,}* ({grant_fraud_rate:.1f}%)\n"
                        f"• Add'l Fraud: *{grant_addl_fraud:,}* ({grant_addl_fraud_rate:.1f}%)\n"
                        f"• Net Valid: *{grant_net_valid:,}*"
            }
        }
    ]
    
    # Add download links
    if kikoff_permalink or grant_permalink:
        download_text = "📎 *Download Reports:*\n"
        if kikoff_permalink:
            download_text += f"• <{kikoff_permalink}|Kikoff Report>\n"
        if grant_permalink:
            download_text += f"• <{grant_permalink}|Grant Report>"
        
        blocks.append({
            "type": "section",
            "text": {
                "type": "mrkdwn",
                "text": download_text
            }
        })
    else:
        github_repo = os.environ.get("GITHUB_REPOSITORY", "")
        github_run_id = os.environ.get("GITHUB_RUN_ID", "")
        
        if github_repo and github_run_id:
            artifact_url = f"https://github.com/{github_repo}/actions/runs/{github_run_id}"
            blocks.append({
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": f"📎 <{artifact_url}|Download Reports> (requires GitHub access)"
                }
            })
    
    message = {"blocks": blocks}
    
    response = requests.post(
        SLACK_WEBHOOK_URL,
        json=message,
        headers={"Content-Type": "application/json"}
    )
    
    if response.status_code == 200:
        print("Slack notification sent successfully")
    else:
        print(f"Slack notification failed: {response.status_code} - {response.text}")


def main():
    """Main execution flow."""
    print("=" * 60)
    print("AppsFlyer Monthly Partner Report")
    print("=" * 60)
    
    # Get date range for previous month
    from_date, to_date = get_previous_month_range()
    report_month = get_report_month_name()
    
    print(f"\nReport Period: {report_month}")
    print(f"Date Range: {from_date} to {to_date}")
    
    # Process both apps
    kikoff_data = process_kikoff_app(from_date, to_date)
    grant_data = process_grant_app(from_date, to_date)
    
    # Generate Excel reports (one per app)
    print("\n" + "-" * 40)
    print("Generating Excel Reports")
    print("-" * 40)
    
    kikoff_filepath = generate_kikoff_excel_report(kikoff_data, report_month)
    grant_filepath = generate_grant_excel_report(grant_data, report_month)
    
    # Send Slack notification
    print("\n" + "-" * 40)
    print("Sending Slack Notification")
    print("-" * 40)
    
    send_combined_slack_notification(kikoff_data, grant_data, report_month, kikoff_filepath, grant_filepath)
    
    print("\n" + "=" * 60)
    print("Report Complete!")
    print("=" * 60)
    
    return kikoff_filepath, grant_filepath


if __name__ == "__main__":
    main()
